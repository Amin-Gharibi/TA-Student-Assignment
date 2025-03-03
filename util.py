from typing import List, Dict
import pandas as pd
import openpyxl
from main import FONT_FILEPATH
import os


def create_excel_output(
        assignment_data: List[Dict[str, str]],
        ta_assignments: Dict[str, List[str]],
) -> None:
    """
    Creates an Excel file with assignment information using Vazirmatn font.
    Results are sorted by date and time.

    Args:
        assignment_data: List of dictionaries with assignment details
        ta_assignments: Dictionary mapping TAs to lists of assigned students
    """
    try:
        # Create a Pandas Excel writer
        writer = pd.ExcelWriter('assignment_results.xlsx', engine='openpyxl')

        # Convert list of dictionaries to DataFrame
        assignments_df = pd.DataFrame(assignment_data)

        # Sort the DataFrame by Date and then by Start Time
        # First, ensure Date is in a sortable format
        assignments_df['SortDate'] = pd.to_datetime(assignments_df['Date'])
        # Parse the start time and combine with date for proper sorting
        assignments_df['SortDateTime'] = pd.to_datetime(
            assignments_df['Date'] + ' ' + assignments_df['Start Time']
        )
        # Sort the DataFrame
        assignments_df = assignments_df.sort_values(['SortDate', 'SortDateTime'])
        # Drop the sorting columns
        assignments_df = assignments_df.drop(['SortDate', 'SortDateTime'], axis=1)

        # Create the full assignments sheet with sorted data
        assignments_df.to_excel(writer, sheet_name='Full Assignments', index=False)

        # Create a private sheet (just student-TA pairings)
        private_df = assignments_df[['Student', 'TA']].copy()
        private_df.to_excel(writer, sheet_name='Private Assignments', index=False)

        # Create a summary sheet
        summary_data = []
        for ta, students in ta_assignments.items():
            summary_data.append({
                "TA": ta,
                "Assigned Students": len(students),
                "Students": ", ".join(students)
            })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Check if the Vazirmatn font file exists
        vazir_font_exists = os.path.exists(FONT_FILEPATH)

        # Color palette inspired by Persian art
        header_bg_color = '1F4E79'  # Rich blue
        header_font_color = 'FFFFFF'  # White
        row_color_1 = 'D9E2F3'  # Light blue
        row_color_2 = 'E6F0FF'  # Very light blue
        border_color = '8EA9DB'  # Medium blue
        highlight_color = 'FFC000'  # Gold for special emphasis

        # Sheet tab colors
        sheet_colors = {
            'Full Assignments': '70AD47',  # Green
            'Private Assignments': '4472C4',  # Blue
            'Summary': 'ED7D31'  # Orange
        }

        # Apply styling to all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # Set sheet tab color
            worksheet.sheet_properties.tabColor = sheet_colors.get(sheet_name, '4472C4')

            # Find the last column with a letter
            from openpyxl.utils import get_column_letter
            last_column = get_column_letter(worksheet.max_column)

            # Create a thick border style
            thick_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='medium', color=border_color),
                right=openpyxl.styles.Side(style='medium', color=border_color),
                top=openpyxl.styles.Side(style='medium', color=border_color),
                bottom=openpyxl.styles.Side(style='medium', color=border_color)
            )

            # Create a thin border style
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin', color=border_color),
                right=openpyxl.styles.Side(style='thin', color=border_color),
                top=openpyxl.styles.Side(style='thin', color=border_color),
                bottom=openpyxl.styles.Side(style='thin', color=border_color)
            )

            # Apply header styling with rich blue background and white text
            header_cells = worksheet['A1':f'{last_column}1']
            for row in header_cells:
                for cell in row:
                    cell.fill = openpyxl.styles.PatternFill(start_color=header_bg_color,
                                                            end_color=header_bg_color,
                                                            fill_type='solid')
                    # Bold header with Vazirmatn font
                    if vazir_font_exists:
                        cell.font = openpyxl.styles.Font(name='Vazirmatn', bold=True, color=header_font_color, size=13)
                    else:
                        # Fallback if font file not found
                        cell.font = openpyxl.styles.Font(bold=True, color=header_font_color, size=13)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    cell.border = thick_border

            # Set row height for the header (adding vertical padding)
            worksheet.row_dimensions[1].height = 35  # Increased header row height for more emphasis

            # Special formatting for the Summary sheet - highlight the TA names
            if sheet_name == 'Summary':
                for row_idx in range(2, worksheet.max_row + 1):
                    ta_cell = worksheet.cell(row=row_idx, column=1)
                    ta_cell.font = openpyxl.styles.Font(name='Vazirmatn' if vazir_font_exists else None,
                                                        bold=True, color='1F4E79', size=12)

            # Center align all cells, apply Vazirmatn font, and add vertical padding to all rows
            # Also apply alternating row colors
            for row_idx in range(2, worksheet.max_row + 1):
                # Set row height (adding vertical padding to each row)
                worksheet.row_dimensions[row_idx].height = 25  # Standard row height with padding

                # Apply alternating row colors
                row_color = row_color_1 if row_idx % 2 == 0 else row_color_2

                # Apply font, alignment, and coloring for each cell in the row
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    # Set background color based on alternating rows
                    cell.fill = openpyxl.styles.PatternFill(start_color=row_color,
                                                            end_color=row_color,
                                                            fill_type='solid')

                    # Apply font
                    font_size = 11
                    if vazir_font_exists:
                        cell.font = openpyxl.styles.Font(name='Vazirmatn', size=font_size)
                    else:
                        cell.font = openpyxl.styles.Font(size=font_size)

                    # Apply alignment
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                    # Apply borders
                    cell.border = thin_border

            # Auto-adjust column width based on content
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # For Persian text, we need a bit more space
                adjusted_width = (max_length + 2) * 1.3
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the Excel file
        writer.close()

        if vazir_font_exists:
            print("Excel file 'assignment_results.xlsx' created successfully with Vazirmatn font.")
        else:
            print(
                "Excel file 'assignment_results.xlsx' created successfully with default font (Vazirmatn font file not found).")

    except Exception as e:
        print(f"Error creating Excel file: {e}")
        # Fallback to text file
        with open("public-result.txt", "w") as public_file:
            # Sort the data before writing to file
            sorted_data = sorted(assignment_data, key=lambda x: (x['Date'], x['Start Time']))
            for item in sorted_data:
                # Modified to include Date and Start Time (no End Time)
                public_file.write(f"{item['Student']}, {item['TA']}, {item['Date']}, {item['Start Time']}\n")
        print("Fallback: Text file created instead.")

    finally:
        # Also sort the data for the private-result.txt file
        sorted_data = sorted(assignment_data, key=lambda x: (x['Date'], x['Start Time']))
        with open("private-result.txt", "w") as private_file:
            for item in sorted_data:
                private_file.write(f"{item['Student']}, {item['TA']}\n")
            print("TXT file 'private-result.txt' created successfully for future usage")
